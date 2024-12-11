from openpyxl import Workbook

# Neue Arbeitsmappe erstellen
wb = Workbook()
ws = wb.active
ws.title = "Verkäufe"

# Überschriften hinzufügen
ws["A1"] = "Produkt"
ws["B1"] = "Preis"
ws["C1"] = "Menge"

# Daten hinzufügen
daten = [
    ("Laptop", 899, 5),
    ("Maus", 29, 12),
    ("Monitor", 299, 8),
    ("Tastatur", 49, 15),
]

for row, (produkt, preis, menge) in enumerate(daten, start=2):
    ws[f"A{row}"] = produkt
    ws[f"B{row}"] = preis
    ws[f"C{row}"] = menge

# Datei speichern
wb.save("e1_verkaufsdaten.xlsx")
