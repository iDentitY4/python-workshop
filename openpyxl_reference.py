from openpyxl import Workbook, load_workbook

# Neues Workbook erstellen
wb = Workbook()
ws = wb.active

# Bestehendes Workbook laden
wb = load_workbook("example.xlsx")


# Neues Arbeitsblatt erstellen
ws1 = wb.create_sheet("Tabelle1", 0)  # An erster Position
ws2 = wb.create_sheet("Tabelle2")  # Am Ende
ws3 = wb.create_sheet("Tabelle3", -1)  # An vorletzter Position

# Arbeitsblatt umbenennen
ws.title = "Neuer Name"

# Arbeitsblätter auflisten
print(wb.sheetnames)


# Wert in Zelle schreiben
ws["A1"] = "Hallo"
ws.cell(row=1, column=1, value="Hallo")

# Wert aus Zelle lesen
wert = ws["A1"].value
wert = ws.cell(row=1, column=1).value


# Zeile hinzufügen
ws.append([1, 2, 3])

# Durch Zeilen iterieren
for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
    print(row)

# Durch Spalten iterieren
for col in ws.iter_cols(min_col=1, max_col=2, values_only=True):
    print(col)


# Einfache Summenformel
ws["A3"] = "=SUM(A1:A2)"

# Durchschnittsberechnung
ws["B3"] = "=AVERAGE(B1:B2)"


from openpyxl.styles import Alignment, Font, PatternFill

# Schriftart ändern
ws["A1"].font = Font(bold=True, size=12)

# Hintergrundfarbe setzen
ws["A1"].fill = PatternFill("solid", fgColor="FFFF00")

# Ausrichtung ändern
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")


# Workbook speichern
wb.save("example_new.xlsx")

# Workbook schließen
wb.close()


# Maximale Zeilen und Spalten ermitteln
max_zeile = ws.max_row
max_spalte = ws.max_column

# Spaltenbuchstaben ermitteln
from openpyxl.utils import get_column_letter

spalte_buchstabe = get_column_letter(1)  # Gibt 'A' zurück
